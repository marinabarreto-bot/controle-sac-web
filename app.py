from flask import Flask, render_template, request, jsonify
from openpyxl import load_workbook

app = Flask(__name__)

ARQUIVO_EXCEL = r"G:\Drives compartilhados\Planilha Rastreio\tb_Fedex.xlsm"
ABA = "DADOS_FEDEX"

COL_AWB = "B"
COL_STATUS = "I"
COL_DESTINO = "C"
COL_DTA_SAIDA = "E"
COL_NUM_PED = "A"
COL_COMENTARIO = "N"


def buscar_dados_awb(awb):

    wb = load_workbook(ARQUIVO_EXCEL, keep_vba=True, data_only=True)
    ws = wb[ABA]

    for row in range(2, ws.max_row + 1):

        awb_planilha = ws[f"{COL_AWB}{row}"].value

        if str(awb_planilha) == awb:

            status = ws[f"{COL_STATUS}{row}"].value
            destino = ws[f"{COL_DESTINO}{row}"].value
            dta_saida = ws[f"{COL_DTA_SAIDA}{row}"].value
            num_ped = ws[f"{COL_NUM_PED}{row}"].value

            if dta_saida:
                try:
                    dta_saida = dta_saida.strftime("%d/%m/%Y")
                except:
                    pass

            wb.close()

            return {
                "status": status,
                "destino": destino,
                "data_saida": dta_saida,
                "pedido": num_ped
            }

    wb.close()
    return None


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/buscar", methods=["POST"])
def buscar():

    awb = request.form["awb"]

    dados = buscar_dados_awb(awb)

    if dados:
        return jsonify(dados)

    return jsonify({"erro": "AWB não encontrado"})


@app.route("/salvar", methods=["POST"])
def salvar():

    awb = request.form["awb"]
    comentario = request.form["comentario"]

    wb = load_workbook(ARQUIVO_EXCEL, keep_vba=True)
    ws = wb[ABA]

    for row in range(2, ws.max_row + 1):

        awb_planilha = ws[f"{COL_AWB}{row}"].value

        if str(awb_planilha) == awb:

            ws[f"{COL_COMENTARIO}{row}"] = comentario
            wb.save(ARQUIVO_EXCEL)
            wb.close()

            return "OK"

    wb.close()
    return "AWB não encontrado"


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)